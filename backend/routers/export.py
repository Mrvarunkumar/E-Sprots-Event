"""
Export router — Excel downloads for admin panel
GET /export/bgmi      → BGMI teams as .xlsx
GET /export/freefire  → Free Fire teams as .xlsx
GET /export/all       → All teams as .xlsx
"""
import io
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from database import admin_db
from auth import get_current_admin
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.units import mm

router = APIRouter(prefix="/export", tags=["Export"])


# ─── STYLE CONSTANTS ─────────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT   = Font(name="Calibri", bold=True, color="FF6B35", size=11)
TITLE_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=14)
TITLE_FILL    = PatternFill("solid", fgColor="0F3460")
EVEN_FILL     = PatternFill("solid", fgColor="16213E")
ODD_FILL      = PatternFill("solid", fgColor="1A1A2E")
ACCENT_FONT   = Font(name="Calibri", color="E94560", bold=True)
NORMAL_FONT   = Font(name="Calibri", color="FFFFFF", size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER   = Border(
    left   = Side(style="thin", color="E94560"),
    right  = Side(style="thin", color="E94560"),
    top    = Side(style="thin", color="E94560"),
    bottom = Side(style="thin", color="E94560"),
)

COLUMNS = [
    ("Team ID",        14),
    ("Game",           10),
    ("Captain Name",   18),
    ("Captain Phone",  14),
    ("Captain USN",    14),
    ("Player 2 Name",  18),
    ("Player 2 Phone", 14),
    ("Player 2 USN",   14),
    ("Player 3 Name",  18),
    ("Player 3 Phone", 14),
    ("Player 3 USN",   14),
    ("Player 4 Name",  18),
    ("Player 4 Phone", 14),
    ("Player 4 USN",   14),
    ("Email",          24),
    ("Branch",         8),
    ("Semester",       9),
    ("Payment Status", 14),
    ("Verified",       10),
    ("Registered At",  20),
]


def _fetch_teams(game: str | None = None) -> list[dict]:
    try:
        query = admin_db.table("teams").select("*").order("created_at", desc=False)
        if game:
            query = query.eq("game", game)
        result = query.execute()
        return result.data or []
    except Exception as e:
        raise HTTPException(500, detail=f"Database error: {e}")


def _build_workbook(teams: list[dict], title: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = title

    # ── Title row ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(COLUMNS))}1")
    title_cell = ws["A1"]
    title_cell.value  = f"🎮 APEX ARENA — {title.upper()} Registrations"
    title_cell.font   = TITLE_FONT
    title_cell.fill   = TITLE_FILL
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 30

    # ── Sub-title row (generated timestamp) ────────────────────────────────
    ws.merge_cells(f"A2:{get_column_letter(len(COLUMNS))}2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Generated: {datetime.now().strftime('%d %b %Y %I:%M %p')}"
    sub_cell.font  = Font(name="Calibri", italic=True, color="AAAAAA", size=9)
    sub_cell.fill  = TITLE_FILL
    sub_cell.alignment = CENTER
    ws.row_dimensions[2].height = 16

    # ── Header row ──────────────────────────────────────────────────────────
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[3].height = 22

    # ── Data rows ───────────────────────────────────────────────────────────
    for row_idx, team in enumerate(teams, start=4):
        row_fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        registered_at = team.get("created_at", "")
        if registered_at:
            try:
                registered_at = datetime.fromisoformat(
                    registered_at.replace("Z", "+00:00")
                ).strftime("%d %b %Y %H:%M")
            except Exception:
                pass

        row_data = [
            team.get("team_id", ""),
            team.get("game", ""),
            team.get("captain_name", ""),
            team.get("captain_phone", ""),
            team.get("captain_usn", ""),
            team.get("player2_name", ""),
            team.get("player2_phone", ""),
            team.get("player2_usn", ""),
            team.get("player3_name", ""),
            team.get("player3_phone", ""),
            team.get("player3_usn", ""),
            team.get("player4_name", ""),
            team.get("player4_phone", ""),
            team.get("player4_usn", ""),
            team.get("email", ""),
            team.get("branch", ""),
            team.get("semester", ""),
            team.get("payment_status", "").upper(),
            "✓ YES" if team.get("is_verified") else "✗ NO",
            registered_at,
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.alignment = CENTER if col_idx in (1, 2, 16, 17, 18, 19) else LEFT
            cell.border    = THIN_BORDER

            if col_idx == 1:  # Team ID — accent color
                cell.font = ACCENT_FONT
            elif col_idx == 19:  # Verified
                cell.font = Font(
                    name="Calibri", color="00FF88" if team.get("is_verified") else "FF4444",
                    bold=True, size=10,
                )
            elif col_idx == 18:  # Payment status
                status_colors = {
                    "VERIFIED": "00FF88",
                    "PAID":     "FFD700",
                    "PENDING":  "FFA500",
                    "REJECTED": "FF4444",
                }
                cell.font = Font(
                    name="Calibri",
                    color=status_colors.get(str(value), "FFFFFF"),
                    bold=True, size=10,
                )
            else:
                cell.font = NORMAL_FONT

        ws.row_dimensions[row_idx].height = 18

    # ── Summary row ─────────────────────────────────────────────────────────
    summary_row = len(teams) + 4
    ws.merge_cells(f"A{summary_row}:{get_column_letter(len(COLUMNS))}{summary_row}")
    s_cell = ws.cell(row=summary_row, column=1)
    s_cell.value = f"Total Teams: {len(teams)}"
    s_cell.font  = Font(name="Calibri", bold=True, color="FF6B35", size=11)
    s_cell.fill  = HEADER_FILL
    s_cell.alignment = CENTER

    # Freeze header rows
    ws.freeze_panes = "A4"

    return wb


def _stream_workbook(wb: Workbook, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

def _build_pdf_table(teams: list[dict], title: str) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        "pdf_title", parent=styles["Title"],
        fontName="Helvetica-Bold", fontSize=16, spaceAfter=18, alignment=1 # Center
    )
    story.append(Paragraph(f"Apex Arena — {title.upper()} Teams SignIn Sheet", title_style))

    # Table Header
    table_data = [
        ["Team ID", "Captain Name", "Branch & Sem", "Player 2", "Player 3", "Player 4", "Signature"]
    ]

    for t in teams:
        # Combine branch and sem
        branch = t.get("branch", "—")
        sem = t.get("semester", "—")
        branch_sem = f"{branch} - Sem {sem}"
        
        table_data.append([
            t.get("team_id", "—"),
            t.get("captain_name", "—"),
            branch_sem,
            t.get("player2_name", "—") or "—",
            t.get("player3_name", "—") or "—",
            t.get("player4_name", "—") or "—",
            "" # Blank for signature
        ])

    # Adjusted column widths for landscape A4 (approx 277mm usable width)
    col_widths = [18*mm, 38*mm, 35*mm, 38*mm, 38*mm, 38*mm, 60*mm]

    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1A1A2E")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor("#FF6B35")),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor("#000000")),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
    ]))

    # Alternating row colors (white/light grey for printing!)
    for i in range(1, len(table_data)):
        bg_color = colors.HexColor("#FFFFFF") if i % 2 == 0 else colors.HexColor("#F5F5F5")
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, i), (-1, i), bg_color),
            ('TEXTCOLOR', (0, i), (-1, i), colors.black),
        ]))

    story.append(t)
    doc.build(story)
    return buf.getvalue()


# ─── GET /{format_type}/{game_path} ─────────────────────────────────────────
@router.get("/{format_type}/{game_path}", summary="Export game registrations")
async def export_game_data(format_type: str, game_path: str, _: dict = Depends(get_current_admin)):
    game_map = {
        "bgmi": "BGMI",
        "freefire": "Free Fire",
        "hackathon": "Hackathon",
        "quiz": "Quiz"
    }
    
    if game_path not in game_map:
        raise HTTPException(404, detail="Game not found")
        
    actual_game = game_map[game_path]
    teams = _fetch_teams(actual_game)
    
    filename_base = actual_game.replace(" ", "")
    timestamp = datetime.now().strftime('%Y%m%d_%H%M')
    
    if format_type == "excel":
        wb = _build_workbook(teams, actual_game)
        filename = f"{filename_base}_Teams_{timestamp}.xlsx"
        return _stream_workbook(wb, filename)
    
    elif format_type == "pdf":
        pdf_data = _build_pdf_table(teams, actual_game)
        filename = f"{filename_base}_Teams_{timestamp}.pdf"
        return StreamingResponse(
            io.BytesIO(pdf_data),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    else:
        raise HTTPException(400, detail="Invalid format. Use 'excel' or 'pdf'")

